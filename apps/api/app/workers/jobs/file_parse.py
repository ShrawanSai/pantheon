from __future__ import annotations

import asyncio
import csv
import io
import os
from collections.abc import Awaitable, Callable
from typing import Any

from sqlalchemy.ext.asyncio import AsyncEngine, AsyncSession, async_sessionmaker, create_async_engine
from supabase import create_client

from apps.api.app.core.config import get_settings
from apps.api.app.db.models import UploadedFile

_worker_engine: AsyncEngine | None = None
_worker_session_factory: async_sessionmaker[AsyncSession] | None = None


def _raw_database_url() -> str:
    # Worker runtime should use pooled DB connections when available.
    # Fall back to DATABASE_URL for local/test environments.
    database_url = os.getenv("DATABASE_POOL_URL") or os.getenv("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_POOL_URL or DATABASE_URL must be set for worker DB sessions.")
    return database_url


def _to_async_driver(dsn: str) -> str:
    if dsn.startswith("postgresql+psycopg://"):
        return dsn
    if dsn.startswith("postgresql://"):
        return dsn.replace("postgresql://", "postgresql+psycopg://", 1)
    if dsn.startswith("sqlite+aiosqlite://"):
        return dsn
    if dsn.startswith("sqlite:///"):
        return dsn.replace("sqlite:///", "sqlite+aiosqlite:///", 1)
    raise RuntimeError("DATABASE_URL must use a PostgreSQL or SQLite DSN.")


def get_worker_engine() -> AsyncEngine:
    global _worker_engine
    if _worker_engine is None:
        _worker_engine = create_async_engine(_to_async_driver(_raw_database_url()), pool_pre_ping=True)
    return _worker_engine


def get_worker_session_factory() -> async_sessionmaker[AsyncSession]:
    global _worker_session_factory
    if _worker_session_factory is None:
        _worker_session_factory = async_sessionmaker(
            bind=get_worker_engine(),
            class_=AsyncSession,
            expire_on_commit=False,
            autoflush=False,
            autocommit=False,
        )
    return _worker_session_factory


async def _download_file_bytes(storage_key: str) -> bytes:
    settings = get_settings()
    client = create_client(settings.supabase_url, settings.supabase_service_role_key)
    data = await asyncio.to_thread(client.storage.from_(settings.supabase_storage_bucket).download, storage_key)
    if isinstance(data, bytes):
        return data
    if isinstance(data, str):
        return data.encode("utf-8")
    raise RuntimeError("Unexpected storage download payload type.")


def _parse_csv(raw_text: str) -> str:
    lines: list[str] = []
    reader = csv.reader(io.StringIO(raw_text))
    for row in reader:
        lines.append(" | ".join(cell.strip() for cell in row))
    return "\n".join(lines).strip()


def _parse_pdf(raw_bytes: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(raw_bytes))
    pages: list[str] = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text.strip())
    return "\n\n".join(pages).strip()


def _parse_docx(raw_bytes: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(raw_bytes))
    paragraphs: list[str] = []
    for para in doc.paragraphs:
        if para.text.strip():
            paragraphs.append(para.text.strip())
    return "\n\n".join(paragraphs).strip()


def _parse_xlsx(raw_bytes: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    sheets: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell).strip() if cell is not None else "" for cell in row]
            rows.append(" | ".join(cells))
        if rows:
            sheets.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sheets).strip()


def _extract_parsed_text(filename: str, raw_bytes: bytes) -> str:
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if extension in {"txt", "md"}:
        return raw_bytes.decode("utf-8", errors="replace")
    if extension == "csv":
        return _parse_csv(raw_bytes.decode("utf-8", errors="replace"))
    if extension == "pdf":
        return _parse_pdf(raw_bytes)
    if extension == "docx":
        return _parse_docx(raw_bytes)
    if extension in {"xlsx", "xls"}:
        return _parse_xlsx(raw_bytes)
    raise ValueError(f"Unsupported file format: {extension or 'unknown'}")


async def file_parse(ctx: dict[str, Any], file_id: str) -> dict[str, Any]:
    session_factory = ctx.get("session_factory") or get_worker_session_factory()
    downloader: Callable[[str], Awaitable[bytes]] = ctx.get("storage_downloader") or _download_file_bytes

    async with session_factory() as session:
        uploaded_file = await session.get(UploadedFile, file_id)
        if uploaded_file is None:
            return {"status": "not_found", "file_id": file_id}

        try:
            file_bytes = await downloader(uploaded_file.storage_key)
            parsed_text = _extract_parsed_text(uploaded_file.filename, file_bytes)
            uploaded_file.parse_status = "completed"
            uploaded_file.parsed_text = parsed_text
            uploaded_file.error_message = None
            await session.commit()
            return {
                "status": "completed",
                "file_id": file_id,
                "parsed_chars": len(parsed_text),
            }
        except Exception as exc:
            uploaded_file.parse_status = "failed"
            uploaded_file.parsed_text = None
            uploaded_file.error_message = str(exc)[:1000]
            await session.commit()
            return {"status": "failed", "file_id": file_id, "error": str(exc)}
